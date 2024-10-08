# -------------------------------------------------------------------------------
# Name:        Scenario Setup Conversion Tool
# Purpose:     Converts a user-provided scenario setup template into required FTOT input files
# -------------------------------------------------------------------------------

import os
import pandas as pd
import openpyxl
from six.moves import input
from shutil import rmtree
from lxml import etree, objectify
from run_upgrade_tool import save_the_new_run_bat_file, write_bat

# Define the location of the template globally
# We reverse out the lib folder from the local tools module path
tools_dir = os.path.split(os.path.realpath(__file__))[0]
ftot_program_directory = os.path.split(tools_dir)[0]
xml_template_file_location = os.path.join(ftot_program_directory, "lib", "v7_temp_Scenario.xml")


# ==============================================================================


def get_user_xlsx_path():
    print("Provide a scenario setup template file to convert to FTOT inputs (drag and drop is fine here):")
    xlsx_path = ""
    while not os.path.exists(xlsx_path):
        xlsx_path = input('----------------------> ').strip('\"')
        print("USER INPUT: the template file: {}".format(xlsx_path))
        if not os.path.exists(xlsx_path):
            print("Path is not valid. Please enter a valid scenario setup template file path.")
    return xlsx_path


# ==============================================================================


def get_scenario_dir():
    # Directory for resulting scenario files
    print("Provide a directory path for generating FTOT input files to be used in the FTOT scenario run (tool will create directory if does not exist): ")
    scenario_dir = ""
    scenario_dir = input('----------------------> ')
    print("USER INPUT: the scenario directory path: {}".format(scenario_dir))

    # If directory already exists, ask user if ok to remove
    if os.path.exists(scenario_dir):
        print("Directory already exists: {}".format(scenario_dir))
        print("Do you want to delete the directory before proceeding? Enter y or n.")
        choice = input('----------------------> ')
        yes_no = False
        while yes_no == False:
            if (choice.lower() == 'y' or choice.lower() == 'yes'):
                yes_no = True
                print("Deleting directory")
                rmtree(scenario_dir)
            elif (choice.lower() == 'n' or choice.lower() == 'no'):
                yes_no = True
            else:
                print("Invalid input")
                print("Do you want to delete the directory before proceeding? Enter y or n.")
                choice = input('----------------------> ')

    if not os.path.exists(scenario_dir):
        try:
            os.makedirs(scenario_dir)
        except:
            print("Could not create scenario directory path. Try again with a valid scenario directory path.")

    return scenario_dir


# ==============================================================================


# Adapted from https://stackoverflow.com/questions/43941365/openpyxl-read-tables-from-existing-data-book-example
def get_table_from_range(ws, tbl):
    # Grab data from table in worksheet
    data = ws[tbl.ref]

    # List all rows including header row
    rows_list = []
    for row in data:
        # List all columns in each row
        cols = []
        for col in row:
            cols.append(col.value)
        rows_list.append(cols)

    # Convert to Pandas DataFrame
    df = pd.DataFrame(data=rows_list[1:], index=None, columns=rows_list[0])

    return df


# ==============================================================================


def create_csv_files(input_data_dir, xlsx_file):
    # Load XLSX workbook
    wb = openpyxl.load_workbook(filename=xlsx_file, read_only=False, keep_vba=False, data_only=True, keep_links=False)

    # Read Configuration tab for default currency
    ws0 = wb['Configuration']
    # Pull default currency for build cost units
    currency_units = ws0['B15'].value

    # Read commodities table from Commodities and Processes tab
    print("Reading in the Commodities and Processes tab")
    ws1 = wb['Commodities and Processes']
    comm_tbl = ws1.tables['commodities']
    comm_data = get_table_from_range(ws1, comm_tbl)
    # Drop blank rows
    comm_data = comm_data.dropna(how='all')
    # Look for duplicates
    if comm_data['Name'].duplicated().any():
        error = ("Error: Commodities table on 'Commodities and Processes' worksheet has duplicate commodity names. Names must be unique.")
        print(error)
        raise IOError(error)
    # Check Name and Phase are fully filled in
    if comm_data['Name'].isna().any() or comm_data['Phase'].isna().any():
        error = ("Error: Missing values in commodities table on 'Commodities and Processes' worksheet. Name and Phase columns must be completely filled out.")
        print(error)
        raise IOError(error)

    # Iteratively read each process into a dictionary of processes
    processes = {}
    cand_gen_flag = False

    for key in ws1.tables:
        if key != 'commodities':
            # Set cand_gen_flag = True if any process has existing facilities N
            tbl = ws1.tables[key]
            top_cell = tbl.ref.split(':')[0]
            process_name = ws1[top_cell].offset(row=-2, column=1).value
            # Check if process_name is empty string, print warning and skip if so
            if not process_name:
                print("Skipping unnamed process")
                continue
            existing_fac = ws1[top_cell].offset(row=-1, column=1).value
            if existing_fac == 'N':
                cand_gen_flag = True

            print("Reading in process {}".format(process_name))
            data = get_table_from_range(ws1, tbl)
            data = data.dropna(how='all')
            if data.shape[0] == 0:
                error = ("Error: Table for process {} on 'Commodities and Processes' worksheet is empty. Process must have at least one input commodity and one output commodity.".format(process_name))
                print(error)
                raise IOError(error)

            proc_input = data[['Inputs', 'Input Amount', 'Input Units']].dropna(how='all')
            # Check Inputs, Input Amount, and Input Units are fully filled in
            if proc_input['Inputs'].isna().any() or proc_input['Input Amount'].isna().any() or proc_input['Input Units'].isna().any():
                error = ("Error: Missing input values in table for process {} on 'Commodities and Processes' worksheet. Inputs, Input Amount, and Input Units columns must be completely filled out.".format(process_name))
                print(error)
                raise IOError(error)
            proc_input = proc_input.assign(io='i')
            proc_input = proc_input.rename(columns={'Inputs': 'commodity', 'Input Amount': 'value', 'Input Units': 'units'})
            proc_output = data[['Outputs', 'Output Amount', 'Output Units']].dropna(how='all')
            # Check Name and Phase are fully filled in
            if proc_output['Outputs'].isna().any() or proc_output['Output Amount'].isna().any() or proc_output['Output Units'].isna().any():
                error = ("Error: Missing output values in table for process {} on 'Commodities and Processes' worksheet. Outputs, Output Amount, and Output Units columns must be completely filled out.".format(process_name))
                print(error)
                raise IOError(error)
            proc_output = proc_output.assign(io='o')
            proc_output = proc_output.rename(columns={'Outputs': 'commodity', 'Output Amount': 'value', 'Output Units': 'units'})
            proc_slate = pd.concat([proc_input, proc_output])

            # Add processor table to processes dictionary, look for duplicates
            if process_name not in processes:
                processes[process_name] = {}
            else:
                error = ("Error: Processes defined on 'Commodities and Processes' worksheet have duplicate process names. Names must be unique.")
                print(error)
                raise IOError(error)
            processes[process_name]['existing_fac'] = existing_fac
            processes[process_name]['proc_slate'] = proc_slate

    # Read facilities table from Facilities and Amounts tab
    print("Reading in the Facilities and Amounts tab")
    ws2 = wb['Facilities and Amounts']
    fac_tbl = ws2.tables['facilities']
    fac_data = get_table_from_range(ws2, fac_tbl)
    fac_data = fac_data.dropna(how='all')
    # Check Facility Type, Facility Name, and Commodity or Process are fully filled in
    if fac_data['Facility Type'].isna().any() or fac_data['Facility Name'].isna().any() or fac_data['Commodity or Process'].isna().any():
        error = ("Error: Missing values in facilities table on 'Facilities and Amounts' worksheet. Facility Type, Facility Name, and Commodity or Process columns must be completely filled out.")
        print(error)
        raise IOError(error)

    schedule_flag = False
    if not fac_data['Schedule'].isna().all():
        schedule_flag = True

    # Write rmp.csv
    rmp_file_path = os.path.join(input_data_dir, 'rmp.csv')
    # Ignore optional inputs: Min Amount, Min Aggregation, Build Cost
    rmp_data = fac_data.loc[fac_data['Facility Type'] == 'raw_material_producer', ['Facility Name', 'Facility Type', 'Commodity or Process',
                                                                                   'Max Amount', 'Units', 'Schedule']]

    # Check that rmp_data has more than zero rows
    if rmp_data.shape[0] == 0:
        error = ("Error: Facilities table on 'Facilities and Amounts' worksheet has no raw material producers. There must be at least one raw material producer.")
        print(error)
        raise IOError(error)
    # Check Max Amount and Units are fully filled in
    if rmp_data['Max Amount'].isna().any() or rmp_data['Units'].isna().any():
        error = ("Error: Missing values for raw material producers in facilities table on 'Facilities and Amounts' worksheet. Max Amount and Units columns must be completely filled out.")
        print(error)
        raise IOError(error)

    rmp_data = pd.merge(rmp_data, comm_data[['Name', 'Phase']],
                        how='left', left_on='Commodity or Process', right_on='Name', indicator=True)
    if sum(rmp_data['_merge'] == 'left_only') > 0:
        error = ("Error: Facility on 'Facilities and Amounts' worksheet references commodity not found in commodities table on 'Commodities and Processes' worksheet.")
        print(error)
        raise IOError(error)
    rmp_data = rmp_data.drop(columns=['Name', '_merge'])
    rmp_data = rmp_data.rename(columns={'Facility Name': 'facility_name', 'Facility Type': 'facility_type',
                                        'Commodity or Process': 'commodity', 'Max Amount': 'value', 'Units': 'units',
                                        'Schedule': 'schedule', 'Phase': 'phase_of_matter'})
    rmp_data = rmp_data.assign(io='o')
    if not comm_data['Max Transport Distance'].isna().all():
        rmp_data = pd.merge(rmp_data, comm_data[['Name', 'Max Transport Distance']],
                            how='left', left_on='commodity', right_on='Name', indicator=True)
        if sum(rmp_data['_merge'] == 'left_only') > 0:
            error = ("Error: Facility on 'Facilities and Amounts' worksheet references commodity not found in commodities table on 'Commodities and Processes' worksheet.")
            print(error)
            raise IOError(error)
        rmp_data = rmp_data.drop(columns=['Name', '_merge'])
        rmp_data = rmp_data.rename(columns={'Max Transport Distance': 'max_transport_distance'})
    # Move 'schedule' column to end or drop if schedules not used
    if schedule_flag:
        rmp_data = rmp_data[[c for c in rmp_data if c not in ['schedule']] + ['schedule']]
    else:
        rmp_data = rmp_data.drop(columns=['schedule'])

    with open(rmp_file_path, "w", newline='') as f:
        print("Writing the rmp.csv file: {}".format(rmp_file_path))
        rmp_data.to_csv(f, index=False)

    # Write proc.csv and proc_cand.csv
    # Always create a proc.csv file (may have no rows)
    # If cand_gen_flag = True, then create a proc_cand.csv
    proc_file_path = os.path.join(input_data_dir, 'proc.csv')
    proc_cand_file_path = os.path.join(input_data_dir, 'proc_cand.csv')

    all_proc_data = fac_data.loc[fac_data['Facility Type'] == 'processor']
    # Look for duplicates
    if all_proc_data['Facility Name'].duplicated().any():
        error = ("Error: Facilities table on 'Facilities and Amounts' worksheet has duplicate facility names for processors. Processor names must be unique.")
        print(error)
        raise IOError(error)

    # Loop through facility rows
    # Append rows processor by processor
    proc_col_names = ['facility_name', 'facility_type', 'commodity', 'value', 'units', 'phase_of_matter', 'io', 'max_capacity', 'min_capacity', 'build_cost', 'schedule']
    proc_cand_col_names = ['facility_name', 'facility_type', 'commodity', 'value', 'units', 'phase_of_matter', 'io', 'schedule']
    proc_rows = pd.DataFrame(columns=proc_col_names)
    proc_cand_rows = pd.DataFrame(columns=proc_cand_col_names)
    for index, row in all_proc_data.iterrows():
        if row['Commodity or Process'] not in processes:
            error = ("Error: Facility on 'Facilities and Amounts' worksheet references process name not found in 'Commodities and Processes' worksheet.")
            print(error)
            raise IOError(error)
        elif processes[row['Commodity or Process']]['existing_fac'] == 'Y':
            print("Adding facility {} to proc.csv file {}".format(row['Facility Name'], proc_file_path))
            # Look up product slate for the process name
            temp_proc = processes[row['Commodity or Process']]['proc_slate']
            temp_proc = pd.merge(temp_proc, comm_data[['Name', 'Phase']],
                                 how='left', left_on='commodity', right_on='Name', indicator=True)
            if sum(temp_proc['_merge'] == 'left_only') > 0:
                error = ("Error: Process {} on 'Commodities and Processes' worksheet references commodity not found in commodities table on 'Commodities and Processes' worksheet.".format(row['Commodity or Process']))
                print(error)
                raise IOError(error)
            temp_proc = temp_proc.drop(columns=['Name', '_merge'])
            temp_proc = temp_proc.rename(columns={'Phase': 'phase_of_matter'})
            # Add 'total' row if needed
            if pd.notnull(row['Max Amount']) or pd.notnull(row['Min Amount']):
                if pd.notnull(row['Units']):
                    # Determine if capacity is in solid or liquid units
                    if row['Units'].lower() in ['tons', 'tonnes', 'pounds', 'kilograms']:
                        cap_phase = 'solid'
                    elif row['Units'].lower() in ['liters', 'gallons', 'thousand gallons']:
                        cap_phase = 'liquid'
                else:
                    error = ("Error: Unit is required in facilities table on 'Facilities and Amounts' worksheet since amount is specified.")
                    print(error)
                    raise IOError(error)
                print("Adding facility capacity row for facility {}".format(row['Facility Name']))
                temp_total = {'commodity': 'total', 'value': '', 'units': row['Units'], 'phase_of_matter': cap_phase, 'io': 'i', 'max_capacity': row['Max Amount'], 'min_capacity': row['Min Amount']}
                temp_proc = pd.concat([temp_proc, pd.DataFrame([temp_total])], ignore_index=True)
            # Add facility_name, facility_type, build cost, schedule fields
            temp_proc = temp_proc.assign(facility_name=row['Facility Name'], facility_type=row['Facility Type'],
                                         build_cost=row['Build Cost'], schedule=row['Schedule'])
            proc_rows = pd.concat([proc_rows, temp_proc])
        elif processes[row['Commodity or Process']]['existing_fac'] == 'N':
            print("Adding candidate processor specification {} to proc_cand.csv file {}".format(row['Facility Name'], proc_cand_file_path))
            if pd.notnull(row['Max Amount']) and pd.notnull(row['Units']) and pd.notnull(row['Min Amount']) and pd.notnull(row['Min Aggregation']) and pd.notnull(row['Build Cost']):
                # Look up product slate for the process name
                temp_proc_cand = processes[row['Commodity or Process']]['proc_slate']
                temp_proc_cand = pd.merge(temp_proc_cand, comm_data[['Name', 'Phase']],
                                          how='left', left_on='commodity', right_on='Name', indicator=True)
                if sum(temp_proc_cand['_merge'] == 'left_only') > 0:
                    error = ("Error: Process {} on 'Commodities and Processes' worksheet references commodity not found in commodities table on 'Commodities and Processes' worksheet.".format(row['Commodity or Process']))
                    print(error)
                    raise IOError(error)
                temp_proc_cand = temp_proc_cand.drop(columns=['Name', '_merge'])
                temp_proc_cand = temp_proc_cand.rename(columns={'Phase': 'phase_of_matter'})
                # Determine if capacity is in solid or liquid units
                if row['Units'].lower() in ['tons', 'tonnes', 'pounds', 'kilograms']:
                    cap_phase = 'solid'
                elif row['Units'].lower() in ['liters', 'gallons', 'thousand gallons']:
                    cap_phase = 'liquid'
                else:
                    error = ("Error: Units not recognized in facilities table on 'Facilities and Amounts' worksheet.")
                    print(error)
                    raise IOError(error)
                # Add maxsize, minsize, min_aggregation, cost_formula rows
                # Use default currency and calculate per unit build cost from row['Build Cost'] based on maxsize
                temp_proc_cand = pd.concat([temp_proc_cand, pd.DataFrame([{'commodity': 'minsize', 'value': row['Min Amount'], 'units': row['Units'], 'phase_of_matter': cap_phase, 'io': ''}])], ignore_index=True)
                temp_proc_cand = pd.concat([temp_proc_cand, pd.DataFrame([{'commodity': 'maxsize', 'value': row['Max Amount'], 'units': row['Units'], 'phase_of_matter': cap_phase, 'io': ''}])], ignore_index=True)
                temp_proc_cand = pd.concat([temp_proc_cand, pd.DataFrame([{'commodity': 'cost_formula', 'value': row['Build Cost']/row['Max Amount'], 'units': currency_units + '/' + row['Units'], 'phase_of_matter': '', 'io': ''}])], ignore_index=True)
                temp_proc_cand = pd.concat([temp_proc_cand, pd.DataFrame([{'commodity': 'min_aggregation', 'value': row['Min Aggregation'], 'units': row['Units'], 'phase_of_matter': cap_phase, 'io': ''}])], ignore_index=True)
                # Add facility_name, facility_type, schedule fields
                temp_proc_cand = temp_proc_cand.assign(facility_name=row['Facility Name'], facility_type=row['Facility Type'], schedule=row['Schedule'])
            else:
                # If Max Amount, Units, Min Amount, Min Aggregation, or Build Cost are blank then give an error
                error = ("Error: Missing values for candidate processors in facilities table on 'Facilities and Amounts' worksheet. Max Amount, Units, Min Amount, Min Aggregation, and Build Cost columns must be completely filled out for candidate processor rows.")
                print(error)
                raise IOError(error)
            proc_cand_rows = pd.concat([proc_cand_rows, temp_proc_cand])
        else:
            error = ("Error: Could not process facilities table on 'Facilities and Amounts' worksheet. Check the contents of the table for faulty inputs.")
            print(error)
            raise IOError(error)

    # Drop build cost, schedule from proc.csv if all null
    if not schedule_flag:
        proc_rows = proc_rows.drop(columns=['schedule'])
        proc_cand_rows = proc_cand_rows.drop(columns=['schedule'])
    if proc_rows['build_cost'].isna().all():
        proc_rows = proc_rows.drop(columns=['build_cost'])

    with open(proc_file_path, "w", newline='') as f:
        print("Writing the proc.csv file: {}".format(proc_file_path))
        proc_rows.to_csv(f, index=False)
    if cand_gen_flag:
        with open(proc_cand_file_path, "w", newline='') as f:
            print("Writing the proc_cand.csv file: {}".format(proc_cand_file_path))
            proc_cand_rows.to_csv(f, index=False)

    # Write dest.csv
    dest_file_path = os.path.join(input_data_dir, 'dest.csv')
    # Ignore optional inputs: Min Amount, Min Aggregation, Build Cost
    dest_data = fac_data.loc[fac_data['Facility Type'] == 'ultimate_destination', ['Facility Name', 'Facility Type', 'Commodity or Process',
                                                                                   'Max Amount', 'Units', 'Schedule']]

    # Check that dest_data has more than zero rows
    if dest_data.shape[0] == 0:
        error = ("Error: Facilities table on 'Facilities and Amounts' worksheet has no ultimate destinations. There must be at least one destination.")
        print(error)
        raise IOError(error)
    # Check Max Amount and Units are fully filled in
    if dest_data['Max Amount'].isna().any() or dest_data['Units'].isna().any():
        error = ("Error: Missing values for ultimate destinations in facilities table on 'Facilities and Amounts' worksheet. Max Amount and Units columns must be completely filled out.")
        print(error)
        raise IOError(error)

    dest_data = pd.merge(dest_data, comm_data[['Name', 'Phase']],
                         how='left', left_on='Commodity or Process', right_on='Name', indicator=True)
    if sum(dest_data['_merge'] == 'left_only') > 0:
        error = ("Error: Facility on 'Facilities and Amounts' worksheet references commodity not found in commodities table on 'Commodities and Processes' worksheet.")
        print(error)
        raise IOError(error)
    dest_data = dest_data.drop(columns=['Name', '_merge'])
    dest_data = dest_data.rename(columns={'Facility Name': 'facility_name', 'Facility Type': 'facility_type',
                                          'Commodity or Process': 'commodity', 'Max Amount': 'value', 'Units': 'units',
                                          'Schedule': 'schedule', 'Phase': 'phase_of_matter'})
    dest_data = dest_data.assign(io='i')
    # Move 'schedule' column to end or drop if schedules not used
    if schedule_flag:
        dest_data = dest_data[[c for c in dest_data if c not in ['schedule']] + ['schedule']]
    else:
        dest_data = dest_data.drop(columns=['schedule'])

    with open(dest_file_path, "w", newline='') as f:
        print("Writing the dest.csv file: {}".format(dest_file_path))
        dest_data.to_csv(f, index=False)

    return cand_gen_flag, schedule_flag


# ==============================================================================


def create_xml_file(scenario_dir, xlsx_file, cand_gen_flag, schedule_flag):
    # Start with the XML schema object
    parser = etree.XMLParser(remove_blank_text=True)
    the_temp_etree = etree.parse(xml_template_file_location, parser)

    # Fill in all cells from XLSX file
    wb = openpyxl.load_workbook(filename=xlsx_file, read_only=False, keep_vba=False, data_only=True, keep_links=False)
    print("Reading in the Configuration tab")
    ws = wb['Configuration']
    # Go cell by cell
    the_temp_etree.find('{Schema_v7.0.0}Scenario_Name').text = ws['B8'].value
    the_temp_etree.find('{Schema_v7.0.0}Scenario_Description').text = ws['B9'].value

    the_temp_etree.find('{Schema_v7.0.0}Scenario_Inputs').find('{Schema_v7.0.0}Common_Data_Folder').text = ws['B32'].value
    the_temp_etree.find('{Schema_v7.0.0}Scenario_Inputs').find('{Schema_v7.0.0}Base_Network_Gdb').text = ws['B33'].value
    the_temp_etree.find('{Schema_v7.0.0}Scenario_Inputs').find('{Schema_v7.0.0}Disruption_Data').text = ws['B37'].value
    the_temp_etree.find('{Schema_v7.0.0}Scenario_Inputs').find('{Schema_v7.0.0}Base_RMP_Layer').text = ws['B18'].value
    the_temp_etree.find('{Schema_v7.0.0}Scenario_Inputs').find('{Schema_v7.0.0}Base_Destination_Layer').text = ws['B20'].value
    the_temp_etree.find('{Schema_v7.0.0}Scenario_Inputs').find('{Schema_v7.0.0}Base_Processors_Layer').text = ws['B19'].value
    the_temp_etree.find('{Schema_v7.0.0}Scenario_Inputs').find('{Schema_v7.0.0}RMP_Commodity_Data').text = os.path.join(scenario_dir, 'input_data', 'rmp.csv')
    the_temp_etree.find('{Schema_v7.0.0}Scenario_Inputs').find('{Schema_v7.0.0}Destinations_Commodity_Data').text = os.path.join(scenario_dir, 'input_data', 'dest.csv')
    # Always create a proc.csv file but only create a proc_cand.csv file if cand_gen_flag is True
    the_temp_etree.find('{Schema_v7.0.0}Scenario_Inputs').find('{Schema_v7.0.0}Processors_Commodity_Data').text = os.path.join(scenario_dir, 'input_data', 'proc.csv')
    if cand_gen_flag:
        the_temp_etree.find('{Schema_v7.0.0}Scenario_Inputs').find('{Schema_v7.0.0}Processors_Candidate_Commodity_Data').text = os.path.join(scenario_dir, 'input_data', 'proc_cand.csv')
    else:
        the_temp_etree.find('{Schema_v7.0.0}Scenario_Inputs').find('{Schema_v7.0.0}Processors_Candidate_Commodity_Data').text = 'None'
    if schedule_flag:
        assert os.path.exists(ws['B38'].value), 'Schedules are specified for facilities in the Facilities and Amounts tab but schedule file {} indicated in cell B38 of the Configuration tab cannot be found.'.format(ws['B38'].value)
    the_temp_etree.find('{Schema_v7.0.0}Scenario_Inputs').find('{Schema_v7.0.0}Schedule_Data').text = ws['B38'].value
    the_temp_etree.find('{Schema_v7.0.0}Scenario_Inputs').find('{Schema_v7.0.0}Commodity_Mode_Data').text = ws['B35'].value
    the_temp_etree.find('{Schema_v7.0.0}Scenario_Inputs').find('{Schema_v7.0.0}Commodity_Density_Data').text = ws['B36'].value
    the_temp_etree.find('{Schema_v7.0.0}Scenario_Inputs').find('{Schema_v7.0.0}Default_Units_Solid_Phase').text = ws['B12'].value
    the_temp_etree.find('{Schema_v7.0.0}Scenario_Inputs').find('{Schema_v7.0.0}Default_Units_Liquid_Phase').text = ('thousand_gallon' if ws['B13'].value == 'thousand gallons' else ws['B13'].value)
    the_temp_etree.find('{Schema_v7.0.0}Scenario_Inputs').find('{Schema_v7.0.0}Default_Units_Distance').text = ws['B14'].value
    the_temp_etree.find('{Schema_v7.0.0}Scenario_Inputs').find('{Schema_v7.0.0}Default_Units_Currency').text = ws['B15'].value

    # Need to convert from "thousand gallons" to "thousand_gallon"
    the_temp_etree.find('{Schema_v7.0.0}Assumptions').find('{Schema_v7.0.0}Truck_Load_Solid').text = str(ws['B54'].value) + ' ' + ws['C54'].value
    the_temp_etree.find('{Schema_v7.0.0}Assumptions').find('{Schema_v7.0.0}Railcar_Load_Solid').text = str(ws['B57'].value) + ' ' + ws['C57'].value
    the_temp_etree.find('{Schema_v7.0.0}Assumptions').find('{Schema_v7.0.0}Barge_Load_Solid').text = str(ws['B60'].value) + ' ' + ws['C60'].value
    the_temp_etree.find('{Schema_v7.0.0}Assumptions').find('{Schema_v7.0.0}Truck_Load_Liquid').text = str(ws['B55'].value) + ' ' + ('thousand_gallon' if ws['C55'].value == 'thousand gallons' else ws['C55'].value)
    the_temp_etree.find('{Schema_v7.0.0}Assumptions').find('{Schema_v7.0.0}Railcar_Load_Liquid').text = str(ws['B58'].value) + ' ' + ('thousand_gallon' if ws['C58'].value == 'thousand gallons' else ws['C58'].value)
    the_temp_etree.find('{Schema_v7.0.0}Assumptions').find('{Schema_v7.0.0}Barge_Load_Liquid').text = str(ws['B61'].value) + ' ' + ('thousand_gallon' if ws['C61'].value == 'thousand gallons' else ws['C61'].value)
    the_temp_etree.find('{Schema_v7.0.0}Assumptions').find('{Schema_v7.0.0}Pipeline_Crude_Load_Liquid').text = str(ws['B63'].value) + ' ' + ('thousand_gallon' if ws['C63'].value == 'thousand gallons' else ws['C63'].value)
    the_temp_etree.find('{Schema_v7.0.0}Assumptions').find('{Schema_v7.0.0}Pipeline_Prod_Load_Liquid').text = str(ws['B64'].value) + ' ' + ('thousand_gallon' if ws['C64'].value == 'thousand gallons' else ws['C64'].value)
    the_temp_etree.find('{Schema_v7.0.0}Assumptions').find('{Schema_v7.0.0}Truck_Fuel_Efficiency').text = str(ws['B56'].value) + ' ' + ws['C56'].value
    the_temp_etree.find('{Schema_v7.0.0}Assumptions').find('{Schema_v7.0.0}Rail_Fuel_Efficiency').text = str(ws['B59'].value) + ' ' + ws['C59'].value
    the_temp_etree.find('{Schema_v7.0.0}Assumptions').find('{Schema_v7.0.0}Barge_Fuel_Efficiency').text = str(ws['B62'].value) + ' ' + ws['C62'].value
    the_temp_etree.find('{Schema_v7.0.0}Assumptions').find('{Schema_v7.0.0}Road_CO2_Emissions').text = str(ws['B67'].value) + ' ' + ws['C67'].value
    the_temp_etree.find('{Schema_v7.0.0}Assumptions').find('{Schema_v7.0.0}Railroad_CO2_Emissions').text = str(ws['B68'].value) + ' ' + ws['C68'].value
    the_temp_etree.find('{Schema_v7.0.0}Assumptions').find('{Schema_v7.0.0}Barge_CO2_Emissions').text = str(ws['B69'].value) + ' ' + ws['C69'].value
    the_temp_etree.find('{Schema_v7.0.0}Assumptions').find('{Schema_v7.0.0}Pipeline_CO2_Emissions').text = str(ws['B70'].value) + ' ' + ws['C70'].value
    the_temp_etree.find('{Schema_v7.0.0}Assumptions').find('{Schema_v7.0.0}Detailed_Emissions_Data').text = ws['B39'].value
    the_temp_etree.find('{Schema_v7.0.0}Assumptions').find('{Schema_v7.0.0}Density_Conversion_Factor').text = str(ws['B71'].value) + ' ' + ws['C71'].value

    the_temp_etree.find('{Schema_v7.0.0}scriptParameters').find('{Schema_v7.0.0}Create_Network_Layer_Script').find('{Schema_v7.0.0}Network_Costs').find('{Schema_v7.0.0}Modal_Costs').find('{Schema_v7.0.0}Truck').find('{Schema_v7.0.0}liquid_Truck_Base_Cost').text = str(ws['B43'].value) + ' ' + ws['C43'].value
    the_temp_etree.find('{Schema_v7.0.0}scriptParameters').find('{Schema_v7.0.0}Create_Network_Layer_Script').find('{Schema_v7.0.0}Network_Costs').find('{Schema_v7.0.0}Modal_Costs').find('{Schema_v7.0.0}Truck').find('{Schema_v7.0.0}solid_Truck_Base_Cost').text = str(ws['B42'].value) + ' ' + ws['C42'].value
    the_temp_etree.find('{Schema_v7.0.0}scriptParameters').find('{Schema_v7.0.0}Create_Network_Layer_Script').find('{Schema_v7.0.0}Network_Costs').find('{Schema_v7.0.0}Modal_Costs').find('{Schema_v7.0.0}Railroad').find('{Schema_v7.0.0}liquid_Railroad_Class_I_Cost').text = str(ws['B45'].value) + ' ' + ws['C45'].value
    the_temp_etree.find('{Schema_v7.0.0}scriptParameters').find('{Schema_v7.0.0}Create_Network_Layer_Script').find('{Schema_v7.0.0}Network_Costs').find('{Schema_v7.0.0}Modal_Costs').find('{Schema_v7.0.0}Railroad').find('{Schema_v7.0.0}solid_Railroad_Class_I_Cost').text = str(ws['B44'].value) + ' ' + ws['C44'].value
    the_temp_etree.find('{Schema_v7.0.0}scriptParameters').find('{Schema_v7.0.0}Create_Network_Layer_Script').find('{Schema_v7.0.0}Network_Costs').find('{Schema_v7.0.0}Modal_Costs').find('{Schema_v7.0.0}Barge').find('{Schema_v7.0.0}liquid_Barge_cost').text = str(ws['B47'].value) + ' ' + ws['C47'].value
    the_temp_etree.find('{Schema_v7.0.0}scriptParameters').find('{Schema_v7.0.0}Create_Network_Layer_Script').find('{Schema_v7.0.0}Network_Costs').find('{Schema_v7.0.0}Modal_Costs').find('{Schema_v7.0.0}Barge').find('{Schema_v7.0.0}solid_Barge_cost').text = str(ws['B46'].value) + ' ' + ws['C46'].value
    the_temp_etree.find('{Schema_v7.0.0}scriptParameters').find('{Schema_v7.0.0}Create_Network_Layer_Script').find('{Schema_v7.0.0}Network_Costs').find('{Schema_v7.0.0}Modal_Costs').find('{Schema_v7.0.0}Artificial_Link').find('{Schema_v7.0.0}liquid_Artificial_Cost').text = str(ws['B49'].value) + ' ' + ws['C49'].value
    the_temp_etree.find('{Schema_v7.0.0}scriptParameters').find('{Schema_v7.0.0}Create_Network_Layer_Script').find('{Schema_v7.0.0}Network_Costs').find('{Schema_v7.0.0}Modal_Costs').find('{Schema_v7.0.0}Artificial_Link').find('{Schema_v7.0.0}solid_Artificial_Cost').text = str(ws['B48'].value) + ' ' + ws['C48'].value
    the_temp_etree.find('{Schema_v7.0.0}scriptParameters').find('{Schema_v7.0.0}Create_Network_Layer_Script').find('{Schema_v7.0.0}Network_Costs').find('{Schema_v7.0.0}Modal_Costs').find('{Schema_v7.0.0}Impedance_Weights_Data').text = ws['B34'].value
    the_temp_etree.find('{Schema_v7.0.0}scriptParameters').find('{Schema_v7.0.0}Create_Network_Layer_Script').find('{Schema_v7.0.0}Network_Costs').find('{Schema_v7.0.0}Intermodal_Transloading_Costs').find('{Schema_v7.0.0}liquid_Transloading_Cost').text = str(ws['B51'].value) + ' ' + ws['C51'].value
    the_temp_etree.find('{Schema_v7.0.0}scriptParameters').find('{Schema_v7.0.0}Create_Network_Layer_Script').find('{Schema_v7.0.0}Network_Costs').find('{Schema_v7.0.0}Intermodal_Transloading_Costs').find('{Schema_v7.0.0}solid_Transloading_Cost').text = str(ws['B50'].value) + ' ' + ws['C50'].value

    the_temp_etree.find('{Schema_v7.0.0}scriptParameters').find('{Schema_v7.0.0}Create_Network_Layer_Script').find('{Schema_v7.0.0}Artificial_Links').find('{Schema_v7.0.0}Road_Max_Artificial_Link_Distance').text = str(ws['B75'].value) + ' ' + ws['C75'].value
    the_temp_etree.find('{Schema_v7.0.0}scriptParameters').find('{Schema_v7.0.0}Create_Network_Layer_Script').find('{Schema_v7.0.0}Artificial_Links').find('{Schema_v7.0.0}Rail_Max_Artificial_Link_Distance').text = str(ws['B76'].value) + ' ' + ws['C76'].value
    the_temp_etree.find('{Schema_v7.0.0}scriptParameters').find('{Schema_v7.0.0}Create_Network_Layer_Script').find('{Schema_v7.0.0}Artificial_Links').find('{Schema_v7.0.0}Water_Max_Artificial_Link_Distance').text = str(ws['B77'].value) + ' ' + ws['C77'].value
    the_temp_etree.find('{Schema_v7.0.0}scriptParameters').find('{Schema_v7.0.0}Create_Network_Layer_Script').find('{Schema_v7.0.0}Artificial_Links').find('{Schema_v7.0.0}Pipeline_Crude_Max_Artificial_Link_Distance').text = str(ws['B78'].value) + ' ' + ws['C78'].value
    the_temp_etree.find('{Schema_v7.0.0}scriptParameters').find('{Schema_v7.0.0}Create_Network_Layer_Script').find('{Schema_v7.0.0}Artificial_Links').find('{Schema_v7.0.0}Pipeline_Products_Max_Artificial_Link_Distance').text = str(ws['B79'].value) + ' ' + ws['C79'].value
    the_temp_etree.find('{Schema_v7.0.0}scriptParameters').find('{Schema_v7.0.0}Create_Network_Layer_Script').find('{Schema_v7.0.0}Artificial_Links').find('{Schema_v7.0.0}Report_With_Artificial_Links').text = str(ws['B74'].value)
    the_temp_etree.find('{Schema_v7.0.0}scriptParameters').find('{Schema_v7.0.0}Create_Network_Layer_Script').find('{Schema_v7.0.0}Short_Haul_Penalties').find('{Schema_v7.0.0}Rail_Short_Haul_Penalty').text = str(ws['B84'].value) + ' ' + ws['C84'].value
    the_temp_etree.find('{Schema_v7.0.0}scriptParameters').find('{Schema_v7.0.0}Create_Network_Layer_Script').find('{Schema_v7.0.0}Short_Haul_Penalties').find('{Schema_v7.0.0}Water_Short_Haul_Penalty').text = str(ws['B85'].value) + ' ' + ws['C85'].value

    the_temp_etree.find('{Schema_v7.0.0}scriptParameters').find('{Schema_v7.0.0}Route_Optimization_Script').find('{Schema_v7.0.0}NDR_On').text = str(ws['B83'].value)
    the_temp_etree.find('{Schema_v7.0.0}scriptParameters').find('{Schema_v7.0.0}Route_Optimization_Script').find('{Schema_v7.0.0}Permitted_Modes').find('{Schema_v7.0.0}Road').text = str(ws['B23'].value)
    the_temp_etree.find('{Schema_v7.0.0}scriptParameters').find('{Schema_v7.0.0}Route_Optimization_Script').find('{Schema_v7.0.0}Permitted_Modes').find('{Schema_v7.0.0}Rail').text = str(ws['B24'].value)
    the_temp_etree.find('{Schema_v7.0.0}scriptParameters').find('{Schema_v7.0.0}Route_Optimization_Script').find('{Schema_v7.0.0}Permitted_Modes').find('{Schema_v7.0.0}Water').text = str(ws['B25'].value)
    the_temp_etree.find('{Schema_v7.0.0}scriptParameters').find('{Schema_v7.0.0}Route_Optimization_Script').find('{Schema_v7.0.0}Permitted_Modes').find('{Schema_v7.0.0}Pipeline_Crude').text = str(ws['B26'].value)
    the_temp_etree.find('{Schema_v7.0.0}scriptParameters').find('{Schema_v7.0.0}Route_Optimization_Script').find('{Schema_v7.0.0}Permitted_Modes').find('{Schema_v7.0.0}Pipeline_Prod').text = str(ws['B27'].value)
    the_temp_etree.find('{Schema_v7.0.0}scriptParameters').find('{Schema_v7.0.0}Route_Optimization_Script').find('{Schema_v7.0.0}Capacity_Options').find('{Schema_v7.0.0}Capacity_On').text = str(ws['B86'].value)
    if str(ws['B86'].value).lower() == 'true':
        assert str(ws['B83'].value).lower() == 'false', 'The NDR_On parameter in cell B83 of the Configuration tab cannot be set to true if the Capacity_On parameter in cell B86 of the Configuration tab is set to true.'
    the_temp_etree.find('{Schema_v7.0.0}scriptParameters').find('{Schema_v7.0.0}Route_Optimization_Script').find('{Schema_v7.0.0}Capacity_Options').find('{Schema_v7.0.0}Background_Flows').find('{Schema_v7.0.0}Road').text = str(ws['B87'].value)
    the_temp_etree.find('{Schema_v7.0.0}scriptParameters').find('{Schema_v7.0.0}Route_Optimization_Script').find('{Schema_v7.0.0}Capacity_Options').find('{Schema_v7.0.0}Background_Flows').find('{Schema_v7.0.0}Rail').text = str(ws['B88'].value)
    the_temp_etree.find('{Schema_v7.0.0}scriptParameters').find('{Schema_v7.0.0}Route_Optimization_Script').find('{Schema_v7.0.0}Capacity_Options').find('{Schema_v7.0.0}Background_Flows').find('{Schema_v7.0.0}Water').text = str(ws['B89'].value)
    the_temp_etree.find('{Schema_v7.0.0}scriptParameters').find('{Schema_v7.0.0}Route_Optimization_Script').find('{Schema_v7.0.0}Capacity_Options').find('{Schema_v7.0.0}Background_Flows').find('{Schema_v7.0.0}Pipeline_Crude').text = str(ws['B90'].value)
    the_temp_etree.find('{Schema_v7.0.0}scriptParameters').find('{Schema_v7.0.0}Route_Optimization_Script').find('{Schema_v7.0.0}Capacity_Options').find('{Schema_v7.0.0}Background_Flows').find('{Schema_v7.0.0}Pipeline_Prod').text = str(ws['B91'].value)
    the_temp_etree.find('{Schema_v7.0.0}scriptParameters').find('{Schema_v7.0.0}Route_Optimization_Script').find('{Schema_v7.0.0}Capacity_Options').find('{Schema_v7.0.0}Minimum_Capacity_Level').text = str(ws['B92'].value)
    the_temp_etree.find('{Schema_v7.0.0}scriptParameters').find('{Schema_v7.0.0}Route_Optimization_Script').find('{Schema_v7.0.0}CO2_Optimization').find('{Schema_v7.0.0}Transport_Cost_Scalar').text = str(ws['B93'].value)
    the_temp_etree.find('{Schema_v7.0.0}scriptParameters').find('{Schema_v7.0.0}Route_Optimization_Script').find('{Schema_v7.0.0}CO2_Optimization').find('{Schema_v7.0.0}CO2_Cost_Scalar').text = str(ws['B94'].value)
    the_temp_etree.find('{Schema_v7.0.0}scriptParameters').find('{Schema_v7.0.0}Route_Optimization_Script').find('{Schema_v7.0.0}CO2_Optimization').find('{Schema_v7.0.0}CO2_Unit_Cost').text = str(ws['B95'].value) + ' ' + ws['C95'].value
    the_temp_etree.find('{Schema_v7.0.0}scriptParameters').find('{Schema_v7.0.0}Route_Optimization_Script').find('{Schema_v7.0.0}Unmet_Demand_Penalty').text = str(ws['B82'].value)

    # Write XML file
    xml_file_path = os.path.join(scenario_dir, "scenario.xml")
    
    if os.path.exists(xml_file_path):
        print(xml_file_path + " already exists")
        print("Do you want to overwrite this XML file? Enter y or n.")
        overwrite = input('----------------------> ')
        yes_no = False
        while yes_no == False:
            if (overwrite.lower() == 'y' or overwrite.lower() == 'yes'):
                yes_no = True
                with open(xml_file_path, 'wb') as wf:
                    print("Writing the XML file: {}".format(xml_file_path))
                    the_temp_etree.write(wf, pretty_print=True)
                    print("Done writing XML file: {}".format(xml_file_path))
                    return xml_file_path
            elif (overwrite.lower() == 'n' or overwrite.lower() == 'no'):
                print("Warning: not overwriting XML file, proceeding to CSV files")
                yes_no = True
            else:
                print("Invalid input")
                print("Do you want to overwrite this XML file? Enter y or n.")
                overwrite = input('----------------------> ')
    else:
        with open(xml_file_path, 'wb') as wf:
            print("Writing the XML file: {}".format(xml_file_path))
            the_temp_etree.write(wf, pretty_print=True)
            print("Done writing XML file: {}".format(xml_file_path))
            return xml_file_path


# ==============================================================================


def create_batch_file(scenario_dir, xml_path, cand_gen_flag):
    # Python environment
    python = "C:\\FTOT\\python3_env\\python.exe"
    print("Default python path: {}".format(python))
    if not os.path.exists(python):
        print("Warning: The default path is not valid. Please enter a valid path to your FTOT python.exe.")
        print("Warning: os.path.exists == False for: {}".format(python))
    while not os.path.exists(python):
        python = input('----------------------> ')
        print("USER INPUT: the python path: {}".format(python))
        if not os.path.exists(python):
            print("Warning: The entered path is not valid. Please enter a valid path to your FTOT python.exe.")
            print("Warning: os.path.exists == False for: {}".format(python))

    # FTOT program directory
    ftot = "C:\\FTOT\\program\\ftot.py"  # Default ftot.py location
    print("Default FTOT path: {}".format(ftot))
    if not os.path.exists(ftot):
        print("Warning: The default path is not valid. Please enter a valid path to your ftot.py file.")
        print("Warning: os.path.exists == False for: {}".format(ftot))
    while not os.path.exists(ftot):
        ftot = input('----------------------> ')
        print("USER INPUT: the ftot.py path: {}".format(ftot))
        if not os.path.exists(ftot):
            print("Warning: The entered path is not valid. Please enter a valid path to your ftot.py file.")
            print("Warning: os.path.exists == False for: {}".format(ftot))

    config_params = [python, ftot, xml_path, cand_gen_flag, scenario_dir]
    # Write XML file
    bat_file_path = save_the_new_run_bat_file(config_params)

    return bat_file_path


# ==============================================================================


def convert_xlsx():

    print("Start: scenario_setup_conversion")

    # Get XLSX file path from user input
    xlsx_path = get_user_xlsx_path()

    # Check scenario setup template contains required tabs
    # Raise error if not
    print("Checking template file contains all required worksheets.")
    wb = openpyxl.load_workbook(filename=xlsx_path, read_only=False, keep_vba=False, data_only=True, keep_links=False)
    required_sheets = ['Configuration', 'Commodities and Processes', 'Facilities and Amounts']
    for sheet in required_sheets:
        if sheet in wb.sheetnames:
            print("Identified worksheet {} for conversion".format(sheet))
        else:
            error = ("Error: missing worksheet {} in template file required for conversion tool".format(sheet))
            print(error)
            raise IOError(error)

    # Get a scenario folder and create an input_data folder
    scenario_dir = get_scenario_dir()
    input_data_dir = os.path.join(scenario_dir, "input_data")
    if not os.path.exists(input_data_dir):
        os.makedirs(input_data_dir)

    # Read in Commodities and Processes tab, Facilities and Amounts tab, create CSV files
    cand_gen_flag, schedule_flag = create_csv_files(input_data_dir, xlsx_path)

    # Read in Configuration tab, create XML file
    xml_path = create_xml_file(scenario_dir, xlsx_path, cand_gen_flag, schedule_flag)

    # Create a batch file based on candidate generation specifications and XML path
    bat_path = create_batch_file(scenario_dir, xml_path, cand_gen_flag)

    print("Finished: scenario_setup_conversion")
        
    return


# ==============================================================================


def run():
    os.system('cls')
    print("FTOT scenario setup conversion tool")
    print("-------------------------------")
    print("")
    print("")
    convert_xlsx()

